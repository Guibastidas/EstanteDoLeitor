"""
HQ Manager - Backend FastAPI
v4.0.0 - Todas as 22 correções aplicadas
"""

import os
import io
import csv
import json
from datetime import datetime
from typing import List, Optional
from contextlib import asynccontextmanager

from fastapi import FastAPI, HTTPException, Query, Depends, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response, StreamingResponse
from pydantic import BaseModel
from sqlalchemy import create_engine, Column, Integer, String, Boolean, Text, ForeignKey, func, case, text
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session, relationship

# ==================== CONFIGURAÇÃO DO BANCO ====================

DATABASE_URL = os.getenv("DATABASE_URL")

if not DATABASE_URL:
    DATABASE_URL = "sqlite:///./hq_manager.db"
    print("⚠️  Usando SQLite local (desenvolvimento)")
    connect_args = {}
else:
    if DATABASE_URL.startswith("postgres://"):
        DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
    print(f"✅ Conectando ao PostgreSQL (Railway)")
    print(f"📊 Database URL: {DATABASE_URL[:20]}...")
    connect_args = {
        "connect_timeout": 10,
        "options": "-c timezone=utc"
    }

try:
    engine = create_engine(
        DATABASE_URL,
        pool_pre_ping=True,
        pool_recycle=300,
        connect_args=connect_args,
        echo=False
    )
    with engine.connect() as conn:
        print("✅ Conexão com banco estabelecida com sucesso!")
except Exception as e:
    print(f"❌ ERRO ao conectar ao banco: {e}")
    raise

SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

# ==================== MODELOS DO BANCO ====================

class SeriesDB(Base):
    __tablename__ = "series"
    id = Column(Integer, primary_key=True, index=True)
    title = Column(String(500), nullable=False, index=True)
    author = Column(String(200))
    publisher = Column(String(200))
    total_issues = Column(Integer, default=0)
    downloaded_issues = Column(Integer, default=0)
    read_issues = Column(Integer, default=0)
    is_completed = Column(Boolean, default=False)
    series_type = Column(String(50), default='em_andamento')
    cover_url = Column(Text)
    notes = Column(Text)
    date_added = Column(String(50), nullable=False)
    date_updated = Column(String(50))
    saga_editions = Column(Text)
    main_issues = Column(Integer, default=0)
    tie_in_issues = Column(Integer, default=0)
    year_start = Column(Integer)
    year_end = Column(Integer)
    issues = relationship("IssueDB", back_populates="series", cascade="all, delete-orphan")


class IssueDB(Base):
    __tablename__ = "issues"
    id = Column(Integer, primary_key=True, index=True)
    series_id = Column(Integer, ForeignKey("series.id", ondelete="CASCADE"), nullable=False)
    issue_number = Column(Integer, nullable=False)
    title = Column(String(500))
    is_read = Column(Boolean, default=False)
    is_downloaded = Column(Boolean, default=True)
    date_added = Column(String(50), nullable=False)
    date_read = Column(String(50))
    series = relationship("SeriesDB", back_populates="issues")


# ==================== MIGRATIONS AUTOMÁTICAS ====================

def run_migrations():
    is_sqlite = DATABASE_URL.startswith("sqlite")
    migrations = [
        ("series", "saga_editions", "TEXT"),
        ("series", "year_start",    "INTEGER"),
        ("series", "year_end",      "INTEGER"),
    ]
    with engine.connect() as conn:
        for table, column, col_type in migrations:
            try:
                if is_sqlite:
                    conn.execute(text(f"ALTER TABLE {table} ADD COLUMN {column} {col_type}"))
                else:
                    conn.execute(text(f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS {column} {col_type}"))
                conn.commit()
                print(f"✅ Migration: coluna '{column}' adicionada à tabela '{table}'")
            except Exception:
                pass


print("📝 Verificando tabelas...")
try:
    Base.metadata.create_all(bind=engine)
    run_migrations()
    print("✅ Tabelas e migrations OK!")
except Exception as e:
    print(f"⚠️ Erro nas migrations: {e}")


# ==================== MODELOS PYDANTIC ====================

class SeriesCreate(BaseModel):
    title: str
    author: Optional[str] = None
    publisher: Optional[str] = None
    total_issues: int = 0
    downloaded_issues: int = 0
    read_issues: int = 0
    is_completed: bool = False
    series_type: str = 'em_andamento'
    cover_url: Optional[str] = None
    notes: Optional[str] = None
    saga_editions: Optional[str] = None
    main_issues: int = 0
    tie_in_issues: int = 0
    year_start: Optional[int] = None
    year_end: Optional[int] = None


class SeriesUpdate(SeriesCreate):
    pass


class SeriesResponse(BaseModel):
    id: int
    title: str
    author: Optional[str]
    publisher: Optional[str]
    total_issues: int
    downloaded_issues: int
    read_issues: int
    is_completed: bool
    series_type: str
    cover_url: Optional[str]
    notes: Optional[str]
    saga_editions: Optional[str]
    status: str
    date_added: str
    date_updated: Optional[str]
    main_issues: int
    tie_in_issues: int
    year_start: Optional[int]
    year_end: Optional[int]

    class Config:
        from_attributes = True


class IssueCreate(BaseModel):
    issue_number: int
    title: Optional[str] = None
    is_read: bool = False


class IssueUpdate(BaseModel):
    is_read: bool


class BulkReadUpdate(BaseModel):
    is_read: bool


class BulkIssueCreate(BaseModel):
    total_issues: int
    read_issues: int = 0
    replace_existing: bool = True


class RangeIssueCreate(BaseModel):
    start_number: int
    end_number: int
    mark_as_read: bool = False


class IssueResponse(BaseModel):
    id: int
    series_id: int
    issue_number: int
    title: Optional[str]
    is_read: bool
    is_downloaded: bool
    date_added: str
    date_read: Optional[str]

    class Config:
        from_attributes = True


# ==================== LIFECYCLE ====================

@asynccontextmanager
async def lifespan(app: FastAPI):
    print("🚀 Iniciando HQ Manager v4.0.0...")
    Base.metadata.create_all(bind=engine)
    run_migrations()
    print("✅ Banco pronto!")
    yield
    print("👋 Encerrando HQ Manager...")


# ==================== APLICAÇÃO ====================

app = FastAPI(
    title="HQ Manager API",
    description="API para gerenciamento de HQs",
    version="4.0.0",
    lifespan=lifespan
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ==================== DEPENDENCY ====================

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ==================== FUNÇÕES AUXILIARES ====================

def calculate_status(read_issues: int, total_issues: int) -> str:
    if read_issues == 0:
        return "para_ler"
    elif read_issues >= total_issues and total_issues > 0:
        return "concluida"
    else:
        return "lendo"


def series_to_response(series: SeriesDB, dl: int = None, rd: int = None) -> dict:
    if dl is not None:
        downloaded_issues = dl if dl > 0 else series.downloaded_issues
        read_issues       = rd if dl > 0 else series.read_issues
    else:
        downloaded_issues = series.downloaded_issues
        read_issues       = series.read_issues

    return {
        "id":               series.id,
        "title":            series.title,
        "author":           series.author,
        "publisher":        series.publisher,
        "total_issues":     series.total_issues,
        "downloaded_issues": downloaded_issues,
        "read_issues":      read_issues,
        "is_completed":     series.is_completed,
        "series_type":      series.series_type,
        "cover_url":        series.cover_url,
        "notes":            series.notes,
        "saga_editions":    series.saga_editions,
        "status":           calculate_status(read_issues, series.total_issues),
        "date_added":       series.date_added,
        "date_updated":     series.date_updated,
        "main_issues":      series.main_issues  or 0,
        "tie_in_issues":    series.tie_in_issues or 0,
        "year_start":       series.year_start,
        "year_end":         series.year_end,
    }


def get_issue_counts(db: Session, series_ids: list) -> dict:
    if not series_ids:
        return {}
    rows = db.query(
        IssueDB.series_id,
        func.count(IssueDB.id).label('dl'),
        func.sum(func.cast(IssueDB.is_read, Integer)).label('rd')
    ).filter(IssueDB.series_id.in_(series_ids)).group_by(IssueDB.series_id).all()
    return {r.series_id: (r.dl or 0, r.rd or 0) for r in rows}


# ==================== SERVIR ARQUIVOS ESTÁTICOS ====================

@app.get("/styles.css")
async def get_styles():
    try:
        with open("styles.css", "r", encoding="utf-8") as f:
            return Response(content=f.read(), media_type="text/css")
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="CSS not found")


@app.get("/script.js")
async def get_script():
    try:
        with open("script.js", "r", encoding="utf-8") as f:
            return Response(content=f.read(), media_type="application/javascript")
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="JS not found")


@app.get("/script-extensions.js")
async def get_script_extensions():
    try:
        with open("script-extensions.js", "r", encoding="utf-8") as f:
            return Response(content=f.read(), media_type="application/javascript")
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="JS extensions not found")


# ==================== APPLE TOUCH ICON (PWA) ====================

@app.get("/apple-touch-icon.png")
@app.get("/apple-touch-icon-precomposed.png")
async def apple_touch_icon():
    """
    Gera o ícone 180x180 para iOS/Safari "Adicionar à Tela de Início".
    Usa Pillow para desenhar o ícone em memória, sem precisar de arquivo estático.
    """
    try:
        from PIL import Image, ImageDraw

        size   = 180
        radius = 38

        img  = Image.new("RGBA", (size, size), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)

        # Fundo âmbar com cantos arredondados
        draw.rounded_rectangle([0, 0, size - 1, size - 1], radius=radius, fill="#d4860a")

        # Livro — três "páginas" empilhadas para dar profundidade
        draw.rounded_rectangle([28, 26, 84, 142], radius=6, fill="#e8dcc8")   # sombra/costas
        draw.rounded_rectangle([36, 26, 92, 142], radius=6, fill="#f5ede0")   # página do meio
        draw.rounded_rectangle([44, 26, 148, 142], radius=6, fill="#ffffff")  # página da frente

        # Lombada
        draw.rectangle([44, 26, 52, 142], fill="#ddd0ba")

        # Título (linha dourada grossa)
        draw.rounded_rectangle([62, 50, 134, 62], radius=4, fill="#d4860a")

        # Linhas de texto
        for i, y in enumerate([76, 88, 100, 112]):
            w = 130 if i % 2 == 0 else 118
            draw.rounded_rectangle([62, y, w, y + 6], radius=3, fill="#c8b89a")

        buf = io.BytesIO()
        img.save(buf, format="PNG", optimize=True)
        buf.seek(0)

        return Response(
            content=buf.read(),
            media_type="image/png",
            headers={"Cache-Control": "public, max-age=86400"}
        )

    except ImportError:
        # Pillow não instalado — retorna 404 com mensagem clara
        raise HTTPException(status_code=404, detail="Pillow não instalado. Adicione 'Pillow' ao requirements.txt")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ENDPOINTS BASE ====================

@app.get("/")
async def root():
    return FileResponse("index.html")


@app.get("/api")
async def api_root():
    return {"message": "HQ Manager API", "version": "4.0.0", "status": "online"}


@app.get("/health")
async def health_check():
    return {"status": "healthy"}


# ==================== SÉRIES ====================

@app.get("/series")
async def list_series(
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(1000, ge=1, le=1000),
    db: Session = Depends(get_db)
):
    try:
        query = db.query(SeriesDB)
        if search:
            pattern = f"%{search}%"
            query = query.filter(
                (SeriesDB.title.ilike(pattern)) |
                (SeriesDB.author.ilike(pattern)) |
                (SeriesDB.publisher.ilike(pattern))
            )
        series_list = query.order_by(SeriesDB.title).all()

        counts = get_issue_counts(db, [s.id for s in series_list])

        result = []
        for s in series_list:
            dl, rd = counts.get(s.id, (0, 0))
            d = series_to_response(s, dl=dl, rd=rd)
            if status and d["status"] != status:
                continue
            result.append(d)

        total_items = len(result)
        if search:
            start = (page - 1) * per_page
            paginated = result[start:start + per_page]
            return {
                "items": paginated,
                "pagination": {
                    "page": page,
                    "per_page": per_page,
                    "total_items": total_items,
                    "total_pages": max(1, -(-total_items // per_page))
                }
            }
        return {
            "items": result,
            "pagination": {
                "page": 1,
                "per_page": total_items,
                "total_items": total_items,
                "total_pages": 1
            }
        }
    except Exception as e:
        print(f"Erro ao listar séries: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/series/{series_id}")
async def get_series(series_id: int, db: Session = Depends(get_db)):
    try:
        series = db.query(SeriesDB).filter(SeriesDB.id == series_id).first()
        if not series:
            raise HTTPException(status_code=404, detail="Série não encontrada")
        counts = get_issue_counts(db, [series_id])
        dl, rd = counts.get(series_id, (0, 0))
        return series_to_response(series, dl=dl, rd=rd)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/series")
async def create_series(series: SeriesCreate, db: Session = Depends(get_db)):
    try:
        db_series = SeriesDB(
            title=series.title,
            author=series.author,
            publisher=series.publisher,
            total_issues=series.total_issues,
            downloaded_issues=series.downloaded_issues,
            read_issues=series.read_issues,
            is_completed=series.is_completed,
            series_type=series.series_type,
            cover_url=series.cover_url,
            notes=series.notes,
            saga_editions=series.saga_editions,
            date_added=datetime.now().isoformat(),
            main_issues=series.main_issues,
            tie_in_issues=series.tie_in_issues,
            year_start=series.year_start,
            year_end=series.year_end,
        )
        db.add(db_series)
        db.commit()
        db.refresh(db_series)
        return series_to_response(db_series)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/series/{series_id}")
async def update_series(series_id: int, series: SeriesUpdate, db: Session = Depends(get_db)):
    try:
        db_s = db.query(SeriesDB).filter(SeriesDB.id == series_id).first()
        if not db_s:
            raise HTTPException(status_code=404, detail="Série não encontrada")

        db_s.title         = series.title
        db_s.author        = series.author
        db_s.publisher     = series.publisher
        db_s.total_issues  = series.total_issues
        db_s.is_completed  = series.is_completed
        db_s.series_type   = series.series_type
        db_s.cover_url     = series.cover_url
        db_s.notes         = series.notes
        db_s.saga_editions = series.saga_editions
        db_s.date_updated  = datetime.now().isoformat()
        db_s.main_issues   = series.main_issues
        db_s.tie_in_issues = series.tie_in_issues
        db_s.year_start    = series.year_start
        db_s.year_end      = series.year_end

        issue_count = db.query(func.count(IssueDB.id)).filter(IssueDB.series_id == series_id).scalar()
        if issue_count == 0:
            db_s.downloaded_issues = series.downloaded_issues
            db_s.read_issues       = series.read_issues

        db.commit()
        db.refresh(db_s)
        counts = get_issue_counts(db, [series_id])
        dl, rd = counts.get(series_id, (0, 0))
        return series_to_response(db_s, dl=dl, rd=rd)
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/series/{series_id}")
async def delete_series(series_id: int, db: Session = Depends(get_db)):
    try:
        db_s = db.query(SeriesDB).filter(SeriesDB.id == series_id).first()
        if not db_s:
            raise HTTPException(status_code=404, detail="Série não encontrada")
        db.delete(db_s)
        db.commit()
        return {"message": "Série deletada com sucesso"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ==================== EDIÇÕES ====================

@app.get("/series/{series_id}/issues")
async def list_issues(series_id: int, db: Session = Depends(get_db)):
    try:
        issues = db.query(IssueDB).filter(
            IssueDB.series_id == series_id
        ).order_by(IssueDB.issue_number).all()
        return [
            {
                "id": i.id, "series_id": i.series_id,
                "issue_number": i.issue_number, "title": i.title,
                "is_read": i.is_read, "is_downloaded": i.is_downloaded,
                "date_added": i.date_added, "date_read": i.date_read
            } for i in issues
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/series/{series_id}/issues")
async def create_issue(series_id: int, issue: IssueCreate, db: Session = Depends(get_db)):
    try:
        series = db.query(SeriesDB).filter(SeriesDB.id == series_id).first()
        if not series:
            raise HTTPException(status_code=404, detail="Série não encontrada")
        now = datetime.now().isoformat()
        db_issue = IssueDB(
            series_id=series_id,
            issue_number=issue.issue_number,
            title=issue.title,
            is_read=issue.is_read,
            is_downloaded=True,
            date_added=now,
            date_read=now if issue.is_read else None
        )
        db.add(db_issue)
        db.commit()
        db.refresh(db_issue)
        return {
            "id": db_issue.id, "series_id": db_issue.series_id,
            "issue_number": db_issue.issue_number, "title": db_issue.title,
            "is_read": db_issue.is_read, "is_downloaded": db_issue.is_downloaded,
            "date_added": db_issue.date_added, "date_read": db_issue.date_read
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/series/{series_id}/issues/bulk")
async def bulk_sync_issues(series_id: int, data: BulkIssueCreate, db: Session = Depends(get_db)):
    try:
        series = db.query(SeriesDB).filter(SeriesDB.id == series_id).first()
        if not series:
            raise HTTPException(status_code=404, detail="Série não encontrada")

        if data.replace_existing:
            db.query(IssueDB).filter(IssueDB.series_id == series_id).delete()

        now = datetime.now().isoformat()
        new_issues = [
            IssueDB(
                series_id=series_id,
                issue_number=i,
                is_read=(i <= data.read_issues),
                is_downloaded=True,
                date_added=now,
                date_read=now if i <= data.read_issues else None
            )
            for i in range(1, data.total_issues + 1)
        ]
        db.bulk_save_objects(new_issues)
        db.commit()
        return {
            "message": f"{data.total_issues} edições sincronizadas",
            "total": data.total_issues,
            "read": data.read_issues
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/series/{series_id}/issues/range")
async def add_issues_by_range(series_id: int, data: RangeIssueCreate, db: Session = Depends(get_db)):
    """Adiciona edições por intervalo (ex: 957 até 1093)"""
    try:
        series = db.query(SeriesDB).filter(SeriesDB.id == series_id).first()
        if not series:
            raise HTTPException(status_code=404, detail="Série não encontrada")
        
        if data.start_number > data.end_number:
            raise HTTPException(status_code=400, detail="Número inicial deve ser menor que o final")
        
        now = datetime.now().isoformat()
        added_count = 0
        
        for i in range(data.start_number, data.end_number + 1):
            # Verifica se já existe
            existing = db.query(IssueDB).filter(
                IssueDB.series_id == series_id,
                IssueDB.issue_number == i
            ).first()
            
            if not existing:
                new_issue = IssueDB(
                    series_id=series_id,
                    issue_number=i,
                    is_read=data.mark_as_read,
                    is_downloaded=True,
                    date_added=now,
                    date_read=now if data.mark_as_read else None
                )
                db.add(new_issue)
                added_count += 1
        
        db.commit()
        return {
            "message": f"{added_count} edições adicionadas (#{data.start_number} a #{data.end_number})",
            "added_count": added_count,
            "range": f"{data.start_number}-{data.end_number}"
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.patch("/series/{series_id}/issues/bulk-read")
async def bulk_update_read(series_id: int, data: BulkReadUpdate, db: Session = Depends(get_db)):
    try:
        now = datetime.now().isoformat() if data.is_read else None
        db.query(IssueDB).filter(IssueDB.series_id == series_id).update(
            {"is_read": data.is_read, "date_read": now},
            synchronize_session=False
        )
        db.commit()
        return {"message": "Edições atualizadas", "is_read": data.is_read}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/series/{series_id}/issues")
async def delete_all_issues(series_id: int, db: Session = Depends(get_db)):
    try:
        series = db.query(SeriesDB).filter(SeriesDB.id == series_id).first()
        if not series:
            raise HTTPException(status_code=404, detail="Série não encontrada")
        deleted = db.query(IssueDB).filter(IssueDB.series_id == series_id).delete()
        db.commit()
        return {"message": f"{deleted} edições deletadas"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/issues/{issue_id}")
async def update_issue(issue_id: int, issue_update: IssueUpdate, db: Session = Depends(get_db)):
    try:
        db_issue = db.query(IssueDB).filter(IssueDB.id == issue_id).first()
        if not db_issue:
            raise HTTPException(status_code=404, detail="Edição não encontrada")
        db_issue.is_read  = issue_update.is_read
        db_issue.date_read = datetime.now().isoformat() if issue_update.is_read else None
        db.commit()
        db.refresh(db_issue)
        return {
            "id": db_issue.id, "series_id": db_issue.series_id,
            "issue_number": db_issue.issue_number, "title": db_issue.title,
            "is_read": db_issue.is_read, "is_downloaded": db_issue.is_downloaded,
            "date_added": db_issue.date_added, "date_read": db_issue.date_read
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.patch("/series/{series_id}/issues/{issue_id}")
async def patch_issue_read_status(
    series_id: int, issue_id: int, update_data: dict, db: Session = Depends(get_db)
):
    try:
        db_issue = db.query(IssueDB).filter(
            IssueDB.id == issue_id, IssueDB.series_id == series_id
        ).first()
        if not db_issue:
            raise HTTPException(status_code=404, detail="Edição não encontrada")
        if "is_read" in update_data:
            db_issue.is_read   = update_data["is_read"]
            db_issue.date_read = datetime.now().isoformat() if update_data["is_read"] else None
        db.commit()
        db.refresh(db_issue)
        return {
            "id": db_issue.id, "series_id": db_issue.series_id,
            "issue_number": db_issue.issue_number, "title": db_issue.title,
            "is_read": db_issue.is_read, "is_downloaded": db_issue.is_downloaded,
            "date_added": db_issue.date_added, "date_read": db_issue.date_read
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/series/{series_id}/issues/{issue_id}")
async def delete_issue(series_id: int, issue_id: int, db: Session = Depends(get_db)):
    try:
        db_issue = db.query(IssueDB).filter(
            IssueDB.id == issue_id, IssueDB.series_id == series_id
        ).first()
        if not db_issue:
            raise HTTPException(status_code=404, detail="Edição não encontrada")
        db.delete(db_issue)
        db.commit()
        return {"message": "Edição deletada com sucesso"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/series/{series_id}/issues/not-downloaded")
async def delete_not_downloaded_issues(series_id: int, db: Session = Depends(get_db)):
    """Deleta todas as edições não baixadas de uma série"""
    try:
        db_series = db.query(SeriesDB).filter(SeriesDB.id == series_id).first()
        if not db_series:
            raise HTTPException(status_code=404, detail="Série não encontrada")
        
        # Buscar e deletar edições não baixadas
        deleted_count = db.query(IssueDB).filter(
            IssueDB.series_id == series_id,
            IssueDB.is_downloaded == False
        ).delete()
        
        db.commit()
        return {
            "message": f"{deleted_count} edições não baixadas foram deletadas",
            "deleted_count": deleted_count
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ESTATÍSTICAS ====================

@app.get("/stats")
async def get_stats(db: Session = Depends(get_db)):
    try:
        total    = db.query(func.count(SeriesDB.id)).scalar() or 0
        para_ler = db.query(func.count(SeriesDB.id)).filter(SeriesDB.read_issues == 0).scalar() or 0
        lendo    = db.query(func.count(SeriesDB.id)).filter(
            SeriesDB.read_issues > 0,
            SeriesDB.read_issues < SeriesDB.total_issues
        ).scalar() or 0
        concluida = db.query(func.count(SeriesDB.id)).filter(
            SeriesDB.read_issues >= SeriesDB.total_issues,
            SeriesDB.total_issues > 0
        ).scalar() or 0
        sagas = db.query(func.count(SeriesDB.id)).filter(
            SeriesDB.series_type == 'saga'
        ).scalar() or 0

        return {
            "total":    total,
            "para_ler": para_ler,
            "lendo":    lendo,
            "concluida": concluida,
            "concluidas": concluida,
            "sagas":    sagas
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== RECALCULAR TUDO ====================

@app.post("/recalculate-all")
async def recalculate_all(db: Session = Depends(get_db)):
    try:
        all_series = db.query(SeriesDB).all()
        counts = get_issue_counts(db, [s.id for s in all_series])
        recalculated_count = 0
        unchanged_count    = 0

        for series in all_series:
            dl, rd = counts.get(series.id, (0, 0))
            if dl > 0:
                if series.downloaded_issues != dl or series.read_issues != rd:
                    series.downloaded_issues = dl
                    series.read_issues       = rd
                    series.date_updated      = datetime.now().isoformat()
                    recalculated_count += 1
            else:
                unchanged_count += 1

        db.commit()
        return {
            "message":      "Recálculo concluído",
            "recalculated": recalculated_count,
            "unchanged":    unchanged_count,
            "total":        len(all_series),
            "errors":       0
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ==================== EXPORTAÇÃO ====================

@app.get("/export")
async def export_data(db: Session = Depends(get_db)):
    try:
        series_list = db.query(SeriesDB).order_by(SeriesDB.title).all()
        counts = get_issue_counts(db, [s.id for s in series_list])

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "id", "title", "author", "publisher", "series_type",
            "total_issues", "downloaded_issues", "read_issues",
            "is_completed", "year_start", "year_end",
            "cover_url", "notes", "saga_editions",
            "main_issues", "tie_in_issues", "date_added", "date_updated"
        ])
        for s in series_list:
            dl, rd = counts.get(s.id, (0, 0))
            downloaded = dl if dl > 0 else s.downloaded_issues
            read       = rd if dl > 0 else s.read_issues
            writer.writerow([
                s.id, s.title, s.author or "", s.publisher or "",
                s.series_type, s.total_issues, downloaded, read,
                s.is_completed, s.year_start or "", s.year_end or "",
                s.cover_url or "", s.notes or "", s.saga_editions or "",
                s.main_issues or 0, s.tie_in_issues or 0,
                s.date_added, s.date_updated or ""
            ])

        output.seek(0)
        filename = f"hq_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== IMPORTAÇÃO ====================

@app.post("/import")
async def import_data(file: UploadFile = File(...), db: Session = Depends(get_db)):
    try:
        content = await file.read()
        filename = file.filename.lower()

        rows = []
        if filename.endswith(".csv"):
            text_content = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text_content))
            rows = list(reader)
        elif filename.endswith((".xlsx", ".xls")):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content))
                ws = wb.active
                headers = [str(cell.value or "").strip() for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
            except ImportError:
                raise HTTPException(status_code=400, detail="openpyxl não disponível para leitura de XLSX")
        else:
            raise HTTPException(status_code=400, detail="Formato inválido. Use CSV ou XLSX.")

        created = 0
        skipped = 0
        errors  = 0
        now     = datetime.now().isoformat()

        for row in rows:
            title = str(row.get("title", "") or row.get("Título", "") or row.get("titulo", "")).strip()
            if not title:
                skipped += 1
                continue
            try:
                def _int(v, default=0):
                    try: return int(v or default)
                    except: return default
                def _bool(v):
                    return str(v).lower() in ("true", "1", "sim", "yes")

                db_s = SeriesDB(
                    title=title,
                    author=str(row.get("author") or row.get("autor") or "").strip() or None,
                    publisher=str(row.get("publisher") or row.get("editora") or "").strip() or None,
                    series_type=str(row.get("series_type") or row.get("tipo") or "em_andamento").strip(),
                    total_issues=_int(row.get("total_issues") or row.get("total")),
                    downloaded_issues=_int(row.get("downloaded_issues") or row.get("baixadas")),
                    read_issues=_int(row.get("read_issues") or row.get("lidas")),
                    is_completed=_bool(row.get("is_completed") or row.get("finalizada")),
                    year_start=_int(row.get("year_start"), None) or None,
                    year_end=_int(row.get("year_end"), None) or None,
                    cover_url=str(row.get("cover_url") or "").strip() or None,
                    notes=str(row.get("notes") or row.get("notas") or "").strip() or None,
                    saga_editions=str(row.get("saga_editions") or "").strip() or None,
                    main_issues=_int(row.get("main_issues")),
                    tie_in_issues=_int(row.get("tie_in_issues")),
                    date_added=now,
                )
                db.add(db_s)
                created += 1
            except Exception as row_err:
                print(f"Erro na linha '{title}': {row_err}")
                errors += 1

        db.commit()
        return {
            "message": f"Importação concluída",
            "created": created,
            "skipped": skipped,
            "errors":  errors
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ==================== RESTORE SAGA ====================

@app.post("/series/{series_id}/restore-saga-values")
async def restore_saga_values(series_id: int, saga_data: dict, db: Session = Depends(get_db)):
    try:
        db_s = db.query(SeriesDB).filter(SeriesDB.id == series_id).first()
        if not db_s:
            raise HTTPException(status_code=404, detail="Série não encontrada")
        if db_s.series_type != 'saga':
            raise HTTPException(status_code=400, detail="Esta série não é uma saga")
        if "main_issues" in saga_data:
            db_s.main_issues = saga_data["main_issues"]
        if "tie_in_issues" in saga_data:
            db_s.tie_in_issues = saga_data["tie_in_issues"]
        db_s.date_updated = datetime.now().isoformat()
        db.commit()
        db.refresh(db_s)
        return series_to_response(db_s)
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ==================== MAIN ====================

if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    print("=" * 70)
    print("HQ MANAGER API v4.0.0")
    print("=" * 70)
    uvicorn.run("main:app", host="0.0.0.0", port=port, reload=False)
